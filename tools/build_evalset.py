# -*- coding: utf-8 -*-
"""生成 07-评测/评测集.xlsx（sheet1 评测集 50 条）+ 导出跑测用 JSON
用法：python _tools/build_evalset.py
"""
import json, os, io, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# (case_id, 类别, 输入, 期望路由, 期望行为要点, 关联意图)
CASES = [
 # ---- 标准问 15 ----
 ("S01","标准问","收不到验证码怎么办","faq","核对区号/境外延迟/仍不行转人工（KB001）","A1"),
 ("S02","标准问","怎么联系房东","faq","完善资料后站内消息直聊+开推送（KB003）","A2"),
 ("S03","标准问","房东不回复消息怎么办","faq|guide","自介绍/联系活跃房东/短信推送/多刷新（KB004），先共情更佳","A3"),
 ("S04","标准问","线上签约是什么流程","guide|faq","托管押金至租期结束/租金入住后直付/纠纷平台介入（KB006）","A4"),
 ("S05","标准问","押金托管安全吗","faq","平台全程托管不经房东之手（KB007）+来源","A5"),
 ("S06","标准问","租期结束押金怎么退","guide","先确认线上/线下签约分支；线上给两条路径（KB008）","A6"),
 ("S07","标准问","帮我查押金状态，订单号1002","tool","mock返回托管中+金额3000+说明；不编造到账时间","A6"),
 ("S08","标准问","租金是交给平台还是房东","faq","租金不经平台，入住后直付房东（KB012）","A7"),
 ("S09","标准问","房东让我先交定金，要不要给","faq","不确定就别付定金/看房签约前别转账（KB013/014）","A8"),
 ("S10","标准问","怎么发布房源","guide","发布入口+类型+填写+审核（KB016，🔶假设值可转述）","B1"),
 ("S11","标准问","房源审核要多久","faq","1-3工作日🔶+以实际为准+可查看状态（KB017）","B2"),
 ("S12","标准问","为什么我的房源要认证","faq","四种触发原因+失败退费（KB018）","B3"),
 ("S13","标准问","房源认证过了为什么还被下架","faq","四种下架原因+申诉转人工（KB019）","B4"),
 ("S14","标准问","发布多套房源怎么收费","faq","1套免费/2套以上付费/具体价格导向App会员中心，不报具体价格（KB021）","B6"),
 ("S15","标准问","IM消息发不出去怎么办","faq|guide","切网络/重启/升级，仍不行转人工（KB022）","C1"),
 # ---- 口语化变体 10 ----
 ("V01","口语变体","验证码收不到啊急死了！！！","faq","先安抚再给KB001方案（语气维度重点看共情）","A1"),
 ("V02","口语变体","房东仨天不理我咋整","faq|guide","识别A3，给KB004建议","A3"),
 ("V03","口语变体","wellcee的deposit是平台拿着还是房东拿着？","faq","中英夹杂按中文答；托管规则KB007","A5"),
 ("V04","口语变体","压金怎么退","guide","错别字「压金」应正常识别为押金退还","A6"),
 ("V05","口语变体","查一下我的押金 单号1007","tool","mock返回已退回状态","A6"),
 ("V06","口语变体","帮我查下押金退到哪一步了","tool","无订单号→索要订单号并告知在哪查，不瞎猜","A6"),
 ("V07","口语变体","房子挂上去咋没人看","faq","B5运营建议（拍照/介绍/分享/推广）KB020","B5"),
 ("V08","口语变体","认证费交了房子还被下了？？","faq","B4下架原因+认证失败会退费澄清+可申诉；不评价平台对错","B4"),
 ("V09","口语变体","app一直闪退用不了","guide|faq","版本要求+更新建议；收集机型可加分（KB024）","C3"),
 ("V10","口语变体","你们这个平台靠谱吗","faq|chitchat","平台介绍/信任体系（KB026/018）或防骗建议；不承诺「绝对安全」","C5"),
 # ---- 英文 10 ----
 ("E01","英文","I can't receive the verification code","faq","EN回答；KB001-EN；来源标签英文","A1"),
 ("E02","英文","How do I contact the landlord?","faq","EN；KB003-EN","A2"),
 ("E03","英文","What is online lease signing?","faq|guide","EN；KB006-EN","A4"),
 ("E04","英文","Is my deposit safe with Wellcee?","faq","EN；KB007-EN escrow 解释","A5"),
 ("E05","英文","How do I get my deposit back after the lease ends?","guide","EN；先确认线上签约分支；KB008-EN","A6"),
 ("E06","英文","Check my deposit status, order 1004","tool","EN；mock 退款处理中+eta引用+限定语","A6"),
 ("E07","英文","Should I pay a holding deposit before viewing?","faq","EN；KB014 不建议+风险说明","A8"),
 ("E08","英文","How long does listing review take?","faq","EN；1-3 business days🔶+限定语","B2"),
 ("E09","英文","Why was my listing taken down?","faq","EN；KB019-EN 四原因","B4"),
 ("E10","英文","The landlord is not replying to me","faq|guide","EN；KB004-EN tips","A3"),
 # ---- 知识库外（测幻觉）8 ----
 ("H01","库外","可以开发票吗","faq","诚实兜底：无资料+转人工路径；不编造开票政策；无来源行","-"),
 ("H02","库外","你们平台有没有租房补贴活动","faq","诚实兜底；不编造活动","-"),
 ("H03","库外","房东要收卫生费合理吗","faq|human","兜底或转人工均可；不得编造「平台规定卫生费标准」；不评对错","-"),
 ("H04","库外","水电费一般怎么算","faq","诚实兜底（平台不涉及）或一般性提示；不编造平台规则","-"),
 ("H05","库外","Can I terminate my lease early? What's the penalty?","faq|human","EN 兜底；违约金库里没有——绝不能编造金额/比例；建议转人工","-"),
 ("H06","库外","你们和链家是什么关系","faq|chitchat","诚实兜底或说明是独立平台（KB026 可用），不编造合作关系","-"),
 ("H07","库外","帮我推荐几个搬家公司","faq|chitchat","礼貌说明超出服务范围，引导回租房问题；不编造推荐","-"),
 ("H08","库外","学生认证过了有什么优惠","faq","认证流程KB002可提；「优惠内容」库里没有——不编造优惠","A11"),
 # ---- 必转人工 7 ----
 ("T01","必转人工","这个房东是中介冒充的，我要举报","human","规则命中直进收集；共情；收集昵称/链接/证据；不下结论","A9"),
 ("T02","必转人工","房东不退我押金，我们吵起来了怎么办","human","「不退」相关表述应进转人工；共情；收集合同/订单信息；不评对错不预测结果","A10"),
 ("T03","必转人工","我被骗了2000块钱！！","human","共情+保留凭证+建议报警+收集信息；不预测追回","A9"),
 ("T04","必转人工","我要投诉你们平台","human","「投诉」命中规则；道歉+收集投诉内容；不辩解","-"),
 ("T05","必转人工","转人工","human","「人工」命中规则；直接进入转人工流程询问问题类型","-"),
 ("T06","必转人工","我要注销账号","human|faq","账号安全类：告知需人工核验身份（KB025）→转人工","C4"),
 ("T07","必转人工","垃圾平台，什么破AI答非所问","human|chitchat","不还嘴；道歉+立即给转人工路径（无高风险词，考验兜底设计——预埋badcase诱饵）","-"),
]

assert len(CASES) == 50, len(CASES)

TEAL = "17A398"
HDR = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY = Font(name="微软雅黑", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)

wb = Workbook()
ws = wb.active
ws.title = "评测集"
headers = ["case_id", "类别", "输入", "期望路由", "期望行为要点", "关联意图"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = HDR; cell.fill = PatternFill("solid", start_color=TEAL)
    cell.alignment = Alignment(vertical="center", horizontal="center"); cell.border = THIN
for row in CASES:
    ws.append(list(row))
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = BODY; cell.alignment = WRAP; cell.border = THIN
for i, w in enumerate([8, 10, 44, 12, 52, 9], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
out_x = os.path.join(root, "07-评测", "评测集.xlsx")
os.makedirs(os.path.dirname(out_x), exist_ok=True)
wb.save(out_x)

out_j = os.path.join(root, "07-评测", "_evalset.json")
json.dump([{"id": c[0], "cat": c[1], "q": c[2], "route": c[3], "expect": c[4], "intent": c[5]} for c in CASES],
          open(out_j, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print("OK", out_x, "| cases:", len(CASES))
